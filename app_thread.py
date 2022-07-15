# -*- coding: utf-8 -*-
# 逻辑线程
# by jhj QQ8510001 2022-7-7
import codecs
import json
import sys
import threading
import time
import warnings
from openpyxl import load_workbook
import traceback

from main import debug

gbfield = ""
gbconst = ""


class AppThread():

    def __init__(self, file, outClientJson, outClientTs, outServerJson, outServerGo, xls_key_id, xls_type_from,
                 xls_rule_from, xls_key_from, xls_content_from):
        super(AppThread, self).__init__()
        self.file_list = file
        self.outClientJson = outClientJson
        self.outClientTs = outClientTs
        self.outServerJson = outServerJson
        self.outServerGo = outServerGo
        # xls 主键
        self.xls_key_id = xls_key_id
        # xls 字段类型开始行
        self.xls_type_from = xls_type_from
        # xls 导出规则开始行
        self.xls_rule_from = xls_rule_from
        # xls 字段开始行
        self.xls_key_from = xls_key_from
        # xls 内容开始行
        self.xls_content_from = xls_content_from

    def run(self):
        try:
            if not self.file_list:
                print("文件为空，线程不工作")
                return

            for f in self.file_list:
                threading.Thread(target=self._gen_json, args=([f])).start()

        except:
            print(sys.exc_info())

    def _gen_json(self, f):

        try:
            warnings.simplefilter("ignore")
            t0 = time.time()

            wb = load_workbook(f[1], read_only=True, data_only=True)
            worksheets = wb.sheetnames
            if worksheets[0] == None:
                return  # 第一个表单没有就终端

            out_file_name = f[0]
            ws = wb[worksheets[0]]

            if ws.max_row == 1 and ws.max_column == 1:
                return  # 没有内容退出

            # 建立存储数据的字典
            # print("当前最大行", ws.max_row, '当前最大列数', ws.max_column)

            # # 因为按行，所以返回A1, B1, C1这样的顺序
            key = {}
            skey = {}
            rule = {}
            type_rule = {}
            data_dict_f = {}
            sdata_dict_f = {}
            i = 0
            pkey = ''
            for row in ws.rows:
                temp_list_f = {}
                stemp_list_f = {}
                i = i + 1
                j = 0
                for cell in row:
                    j += 1
                    if i == 1 and j == 1:  # 主KEY
                        if cell.value is not None:
                            pkey = cell.value
                        else:
                            debug("[错误-没有主键]" + out_file_name + " 行-列-值", i, j, cell.value)
                            break

                    if i == self.xls_type_from:  # type

                        if cell.value is not None:
                            # print("type_rule规则", cell.value, j)
                            type_rule[j] = str(cell.value).strip()

                        else:
                            continue

                    if i == self.xls_rule_from:  # rule
                        if cell.value is not None:
                            # print("前端 后端 或略规则", cell.value, j)
                            rule[j] = str(cell.value).strip()

                        else:
                            continue

                    if i == self.xls_key_from:  # key

                        if cell.value is not None:
                            if str(rule[j]).lower() == "client" or str(rule[j]).lower() == "both":
                                # 容错判断下这个键位的规则有没有
                                if j in type_rule:
                                    # print("导出Client", cell.value, j, rule[j])
                                    key[j] = str(cell.value).strip()
                                else:
                                    debug("[错误--定义了client或both 却没有定义 键或规则 行 - 列]", i, j, out_file_name)

                            if str(rule[j]).lower() == "server" or str(rule[j]).lower() == "both":
                                # 容错判断下这个键位的规则有没有
                                if j in type_rule:
                                    # print("导出Server", cell.value, j, rule[j])
                                    skey[j] = str(cell.value).strip()
                                else:
                                    debug("[错误--定义了client或both 却没有定义 键或规则 行 - 列]", i, j, out_file_name)


                        else:
                            break

                    if i >= self.xls_content_from:  # 内容
                        if j <= len(rule):
                            if j in key:
                                if type_rule[j] == 'int':
                                    if type(cell.value) == str:
                                        if len(cell.value) == 0 or len(str(cell.value).strip()) == 0:
                                            temp_list_f[key[j]] = 0
                                        else:
                                            temp_list_f[key[j]] = round(int(str(cell.value).strip()))
                                    elif cell.value is None:
                                        temp_list_f[key[j]] = 0
                                    else:
                                        temp_list_f[key[j]] = round(int(cell.value))

                                elif type_rule[j] == 'string':
                                    temp_list_f[key[j]] = str(cell.value)
                                elif type_rule[j] == 'bool':
                                    temp_list_f[key[j]] = bool(cell.value)
                                else:
                                    debug("[警告-前端配置值错误]" + out_file_name + " 行-列-类型-值", i, j, type_rule[j], cell.value)

                            if j in skey:
                                if type_rule[j] == 'int':
                                    if type(cell.value) == str:
                                        if len(cell.value) == 0 or len(str(cell.value).strip()) == 0:
                                            stemp_list_f[skey[j]] = 0
                                        else:
                                            stemp_list_f[skey[j]] = round(int(str(cell.value).strip()))
                                    elif cell.value is None:
                                        stemp_list_f[skey[j]] = 0
                                    else:
                                        stemp_list_f[skey[j]] = round(int(cell.value))

                                elif type_rule[j] == 'string':
                                    stemp_list_f[skey[j]] = str(cell.value)
                                elif type_rule[j] == 'bool':
                                    stemp_list_f[skey[j]] = bool(cell.value)
                                else:
                                    debug("[警告-后端配置值错误]" + out_file_name + " 行-列-类型-值", i, j, type_rule[j], cell.value)

                if temp_list_f != {}:
                    if self.xls_key_id not in temp_list_f:
                        temp_list_f[self.xls_key_id] = ''  # 定义一个空的
                        # 判断下主键是多个还是单个
                        if len(pkey.split(',')) > 1:
                            for k in pkey.split(','):
                                for dkey in temp_list_f:
                                    if dkey == k:
                                        temp_list_f[self.xls_key_id] += str(temp_list_f[dkey])
                        else:
                            temp_list_f[self.xls_key_id] = str(temp_list_f[pkey])

                    data_dict_f[temp_list_f[self.xls_key_id]] = temp_list_f

                if stemp_list_f != {}:
                    if self.xls_key_id not in stemp_list_f:
                        stemp_list_f[self.xls_key_id] = ''  # 定义一个空的
                        # 判断下主键是多个还是单个
                        if len(pkey.split(',')) > 1:
                            for k in pkey.split(','):
                                for dkey in stemp_list_f:
                                    if dkey == k:
                                        stemp_list_f[self.xls_key_id] += str(stemp_list_f[dkey])
                        else:
                            stemp_list_f[self.xls_key_id] = str(stemp_list_f[pkey])

                    sdata_dict_f[stemp_list_f[self.xls_key_id]] = stemp_list_f

            if data_dict_f:
                # client json
                self.write_file(self.outClientJson + "/" + out_file_name + ".json", data_dict_f)
                # client ts
                # print(key)
                # print(type_rule)
                self.gents(out_file_name, key, type_rule)

                del temp_list_f
                del data_dict_f
            if sdata_dict_f:
                # server json
                self.write_file(self.outServerJson + "/" + out_file_name + ".json", sdata_dict_f)
                # server go
                self.gengo(out_file_name, skey, type_rule)

                self.genConfigGo(out_file_name)

                del stemp_list_f
                del sdata_dict_f

            t1 = time.time()
            print("[导出完毕]" + out_file_name + "  行：", ws.max_row, '列：', ws.max_column, " 耗时：" + str(t1 - t0))

        except:
            debug("[错误][配置表错误]", f[0], sys.exc_info())
            # debug("[错误][配置表错误]", f[0], "键", key)
            # debug("[错误][配置表错误]", f[0], "类型", type_rule)
            traceback.print_tb(sys.exc_info()[2])

    def write_file(self, filename, buf):
        totxt = codecs.open(filename, 'w', "utf-8")
        totxt.write(str(json.dumps(buf, ensure_ascii=False)))
        totxt.close()

    def gents(self, file, key, type_rule):
        version = '3.4'

        # 生成TS文件
        ts_waring = '/**\n * @导出自 ' + file + '.xls\n * @此文件为自动导出 请勿修改\n * @导出时间 ' + time.strftime('%Y.%m.%d',
                                                                                                  time.localtime(
                                                                                                      time.time())) + '\n * @导出工具 v' + version + ' \n * @Author jhj \n * @QQ 8510001 \n */ \n'
        # 字段
        field = ""
        for k in key:
            if key[k] != 'id':
                field += '\n    get ' + self.toSmaillWord(key[k]) + '(): ' + self.changetype(
                    type_rule[k]) + ' { \n        return this.data.' + key[k] + ';\n    }\n'

        script = ts_waring + 'import {JsonUtil} from  "../../../core/utils/JsonUtil"; \n\n' \
                 + 'export class Table' + self.toBigWord(file) + ' {\n' \
                 + '    static TableName: string = "' + file + '";\n' \
                 + '    private data: any;\n\n' \
                 + '    init(id: number) {\n' \
                 + '        let table = JsonUtil.get(Table' + self.toBigWord(file) + '.TableName);\n' \
                 + '        this.data = table[id];\n' \
                 + '        this.id = id;\n    }\n\n' \
                 + '    id: number = 0;\n' \
                 + field + '\n}'

        totxt_b = codecs.open(self.outClientTs + '/' + file + '.ts', 'w', "utf-8")
        totxt_b.write(script)
        totxt_b.close()

    def changetype(self, type):
        if type.lower() == 'int':
            return 'number'
        else:
            return type

    def changeInt(self, type):
        if type.lower() == 'int':
            return 'uint64'
        else:
            return type

    def gengo(self, file, key, type_rule):
        version = '3.4'

        # 生成TS文件
        ts_waring = '/**\n * @导出自 ' + file + '.xls\n * @此文件为自动导出 请勿修改\n * @导出时间 ' + time.strftime('%Y.%m.%d',
                                                                                                  time.localtime(
                                                                                                      time.time())) + '\n * @导出工具 v' + version + ' \n * @Author jhj \n * @QQ 8510001 \n */ \n'
        # 字段
        field = ""
        for k in key:
            # print(key[k],type_rule[k])
            field += '	' + self.toBigWord(key[k]) + '  ' + self.changeInt(type_rule[k]) + '  `json:"' + key[
                k] + '"`\n'

        script = ts_waring + 'package configdef \n\ntype ' + self.toBigWord(
            file) + ' struct {\n' + field + '}\n\nvar ' + self.toBigWord(file) + 'M  map[uint64]*' + self.toBigWord(
            file)
        totxt_b = codecs.open(self.outServerGo + '/' + file + '.go', 'w', "utf-8")
        totxt_b.write(script)
        totxt_b.close()

    # 下划线转大驼峰
    def toBigWord(self, word):
        out = ""
        if "_" in word:
            for w in word.split("_"):
                out += w.capitalize()

        else:
            return word.capitalize()

        return out

    # 下划线转小驼峰
    def toSmaillWord(self, word):
        out = ""
        if "_" in word:
            for w in word.split("_"):
                if w != word.split("_")[0]:
                    out += w.capitalize()
                else:
                    out += w

        else:
            return word

        return out

    def genConfigGo(self, fname):
        version = '3.4'

        # 生成TS文件
        ts_waring = '/**\n * @此文件为自动导出 请勿修改\n * @导出时间 ' + time.strftime('%Y.%m.%d', time.localtime(
            time.time())) + '\n * @导出工具 v' + version + ' \n * @Author jhj \n * @QQ 8510001 \n */ \n'

        # 字段
        global gbfield, gbconst
        # field = ""
        # for k in filelist:
        #     # print(k[0])
        gbfield += '	case File' + self.toBigWord(fname) + ':\n		return &' + self.toBigWord(fname) + 'M\n'

        # const = ""
        # for k in filelist:
        gbconst += '	File' + self.toBigWord(fname) + ' string = "' + fname + '.json"  \n'

        script = ts_waring + 'package configdef \n\nconst (\n   ' + gbconst + ')\nfunc LoadStrut(fileName string) interface {}  {\n	switch fileName {\n  ' + gbfield + '	default:\n		return nil\n	}\n    return nil\n}'
        totxt_b = codecs.open(self.outServerGo + '/configdef.go', 'w', "utf-8")
        totxt_b.write(script)
        totxt_b.close()
